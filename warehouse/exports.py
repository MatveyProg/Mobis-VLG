from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from warehouse.models import PurchaseRequisition


def export_requisition_xlsx(requisition: PurchaseRequisition) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заявка'

    ws.append([f'Заявка на закупку {requisition.number}'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Статус', requisition.get_status_display()])
    ws.append(['Дата', requisition.created_at.strftime('%d.%m.%Y %H:%M')])
    if requisition.note:
        ws.append(['Комментарий', requisition.note])
    ws.append([])

    headers = [
        '№ п/п',
        'Артикул',
        'Количество',
        'Наименование (англ)',
        'Цена ₩/ед',
        'Сумма ₩',
        'Цена ₽/ед',
        'Сумма ₽',
    ]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    total_krw = Decimal('0')
    total_rub = Decimal('0')
    for item in requisition.items.select_related('product'):
        sum_krw = item.sum_krw
        sum_rub = item.sum_rub
        total_krw += sum_krw
        total_rub += sum_rub
        ws.append([
            item.line_number,
            item.product.sku,
            item.quantity,
            item.product.name_en or item.product.name_ru,
            float(item.purchase_price_krw),
            float(sum_krw),
            float(item.purchase_price_rub),
            float(sum_rub),
        ])

    ws.append([])
    ws.append(['', '', '', '', 'Итого ₩:', float(total_krw), 'Итого ₽:', float(total_rub)])
    for cell in ws[ws.max_row]:
        if cell.value:
            cell.font = Font(bold=True)

    filename = f'requisition_{requisition.number}.xlsx'.replace('/', '-')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
