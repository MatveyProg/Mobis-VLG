from datetime import datetime, time
from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook

from catalog.models import Product
from orders.models import Order, OrderItem
from warehouse.models import StockInItem, StockMovement


def _parse_dates(request):
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    tz = timezone.get_current_timezone()
    start = end = None
    if date_from:
        start = timezone.make_aware(datetime.combine(datetime.strptime(date_from, '%Y-%m-%d').date(), time.min), tz)
    if date_to:
        end = timezone.make_aware(datetime.combine(datetime.strptime(date_to, '%Y-%m-%d').date(), time.max), tz)
    return start, end, date_from or '', date_to or ''


def _xlsx_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def reports_index(request):
    return render(request, 'reports/index.html')


@staff_member_required
def report_stock_in(request):
    start, end, date_from, date_to = _parse_dates(request)
    supplier_id = request.GET.get('supplier')
    qs = StockInItem.objects.filter(stock_in__is_posted=True).select_related(
        'stock_in__supplier', 'product'
    )
    if start:
        qs = qs.filter(stock_in__date__gte=start)
    if end:
        qs = qs.filter(stock_in__date__lte=end)
    if supplier_id:
        qs = qs.filter(stock_in__supplier_id=supplier_id)

    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Приход'
        ws.append(['Дата', 'Поставщик', 'Артикул', 'Наименование', 'Кол-во', 'Цена закупки ₽', 'Сумма ₽'])
        for row in qs:
            ws.append([
                timezone.localtime(row.stock_in.date).strftime('%d.%m.%Y %H:%M'),
                row.stock_in.supplier.name,
                row.product.sku,
                row.product.name_ru,
                row.quantity,
                float(row.purchase_price_rub),
                float(row.quantity * row.purchase_price_rub),
            ])
        return _xlsx_response(wb, 'stock_in.xlsx')

    from warehouse.models import Supplier
    return render(request, 'reports/stock_in.html', {
        'rows': qs[:500],
        'date_from': date_from,
        'date_to': date_to,
        'suppliers': Supplier.objects.all(),
        'supplier_id': supplier_id or '',
    })


@staff_member_required
def report_stock_out(request):
    start, end, date_from, date_to = _parse_dates(request)
    qs = OrderItem.objects.filter(
        order__status__in=[
            Order.Status.CONFIRMED,
            Order.Status.ASSEMBLED,
            Order.Status.SHIPPED,
        ]
    ).select_related('order', 'order__customer', 'product')
    if start:
        qs = qs.filter(order__status_changed_at__gte=start)
    if end:
        qs = qs.filter(order__status_changed_at__lte=end)

    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Расход'
        ws.append(['Дата', '№ заказа', 'Клиент', 'Артикул', 'Наименование', 'Кол-во', 'Цена ₽', 'Сумма ₽'])
        for row in qs:
            ws.append([
                timezone.localtime(row.order.status_changed_at).strftime('%d.%m.%Y %H:%M'),
                row.order.number,
                row.order.customer_name,
                row.product.sku,
                row.product.name_ru,
                row.quantity,
                float(row.price_rub),
                float(row.line_total),
            ])
        return _xlsx_response(wb, 'stock_out.xlsx')

    return render(request, 'reports/stock_out.html', {
        'rows': qs[:500],
        'date_from': date_from,
        'date_to': date_to,
    })


def _avg_purchase_price(product_id):
    items = StockInItem.objects.filter(
        product_id=product_id,
        stock_in__is_posted=True,
    ).values_list('quantity', 'purchase_price_rub')
    total_qty = 0
    total_sum = Decimal('0')
    for qty, price in items:
        total_qty += qty
        total_sum += Decimal(qty) * price
    if total_qty == 0:
        return Decimal('0')
    return (total_sum / total_qty).quantize(Decimal('0.01'))


@staff_member_required
def report_balances(request):
    products = Product.objects.select_related('stock').filter(is_active=True)
    rows = []
    for p in products:
        stock = getattr(p, 'stock', None)
        on_hand = stock.quantity_on_hand if stock else 0
        reserved = stock.quantity_reserved if stock else 0
        available = stock.quantity_available if stock else 0
        avg = _avg_purchase_price(p.id)
        rows.append({
            'sku': p.sku,
            'name': p.name_ru,
            'on_hand': on_hand,
            'reserved': reserved,
            'available': available,
            'avg_purchase': avg,
            'cost_purchase': avg * on_hand,
            'price_rub': p.price_rub,
            'cost_sale': p.price_rub * on_hand,
        })

    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Остатки'
        ws.append([
            'Артикул', 'Наименование', 'На складе', 'Резерв', 'Доступно',
            'Закупка ₽/ед', 'Стоимость закупки', 'Продажа ₽/ед', 'Стоимость продажи',
        ])
        for r in rows:
            ws.append([
                r['sku'], r['name'], r['on_hand'], r['reserved'], r['available'],
                float(r['avg_purchase']), float(r['cost_purchase']),
                float(r['price_rub']), float(r['cost_sale']),
            ])
        return _xlsx_response(wb, 'balances.xlsx')

    return render(request, 'reports/balances.html', {'rows': rows})


@staff_member_required
def report_movement(request):
    start, end, date_from, date_to = _parse_dates(request)
    products = Product.objects.select_related('stock').all()
    rows = []
    for p in products:
        movements = StockMovement.objects.filter(product=p)
        if start:
            before = movements.filter(created_at__lt=start)
            period = movements.filter(created_at__gte=start)
        else:
            before = StockMovement.objects.none()
            period = movements
        if end:
            period = period.filter(created_at__lte=end)

        def net(qs):
            income = qs.filter(movement_type__in=['IN', 'STORNO_OUT']).aggregate(s=Sum('quantity'))['s'] or 0
            outcome = qs.filter(movement_type='OUT').aggregate(s=Sum('quantity'))['s'] or 0
            return income - outcome

        opening = net(before)
        income = period.filter(movement_type='IN').aggregate(s=Sum('quantity'))['s'] or 0
        outcome = period.filter(movement_type='OUT').aggregate(s=Sum('quantity'))['s'] or 0
        storno = period.filter(movement_type='STORNO_OUT').aggregate(s=Sum('quantity'))['s'] or 0
        closing = opening + income - outcome + storno
        reserved = p.stock.quantity_reserved if hasattr(p, 'stock') else 0
        if opening or income or outcome or storno or closing:
            rows.append({
                'sku': p.sku,
                'name': p.name_ru,
                'opening': opening,
                'income': income,
                'outcome': outcome,
                'storno': storno,
                'closing': closing,
                'reserved': reserved,
            })

    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Движение'
        ws.append(['Артикул', 'Наименование', 'Начало', 'Приход', 'Расход', 'Сторно', 'Конец', 'Резерв'])
        for r in rows:
            ws.append([r['sku'], r['name'], r['opening'], r['income'], r['outcome'], r['storno'], r['closing'], r['reserved']])
        return _xlsx_response(wb, 'movement.xlsx')

    return render(request, 'reports/movement.html', {
        'rows': rows,
        'date_from': date_from,
        'date_to': date_to,
    })
